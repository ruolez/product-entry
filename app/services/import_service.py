import json as _json
import os
import uuid
import tempfile
import threading
from io import BytesIO

from openpyxl import load_workbook

from models.settings import db
from services.store_connection import execute_query, get_store
from services.lookup_service import get_all_subcategories
from services.item_service import insert_item, get_field_configs


UPLOAD_DIR = tempfile.gettempdir()
_file_cache = {}
_progress_lock = threading.Lock()

COLUMN_ALIASES = {
    "ProductUPC": [
        "upc", "barcode", "productupc", "product upc", "item upc",
        "upc code", "upc/barcode", "product barcode",
    ],
    "ProductSKU": [
        "sku", "productsku", "product sku", "item sku", "sku code",
        "item code", "item number", "item #", "item no",
    ],
    "ProductDescription": [
        "description", "name", "product name", "productdescription",
        "product description", "item name", "item description", "product",
    ],
    "SubCateName": [
        "subcategory", "sub category", "subcatename", "sub category name",
        "subcategory name", "sub-category", "category", "main category",
        "categoryname", "category name", "dept", "department",
    ],
    "UnitCost": ["unit cost", "cost", "unitcost", "cost price"],
    "UnitPrice": [
        "unit price", "price", "unitprice", "standard price",
        "retail price", "retail", "sell price", "selling price",
    ],
    "UnitPriceC": [
        "unitpricec", "price c", "delivery b", "delivery b price",
        "del b price", "pricec",
    ],
}


def store_upload(file_bytes):
    upload_id = str(uuid.uuid4())
    path = os.path.join(UPLOAD_DIR, f"import_{upload_id}.xlsx")
    with open(path, "wb") as f:
        f.write(file_bytes)
    _file_cache[upload_id] = {"path": path}
    return upload_id


def get_upload(upload_id):
    entry = _file_cache.get(upload_id)
    if entry and os.path.exists(entry["path"]):
        with open(entry["path"], "rb") as f:
            return f.read()
    # Fallback: look on disk (handles multi-worker / restart scenarios)
    path = os.path.join(UPLOAD_DIR, f"import_{upload_id}.xlsx")
    if os.path.exists(path):
        _file_cache[upload_id] = {"path": path}
        with open(path, "rb") as f:
            return f.read()
    return None


def cleanup_upload(upload_id):
    entry = _file_cache.pop(upload_id, None)
    path = entry["path"] if entry else os.path.join(UPLOAD_DIR, f"import_{upload_id}.xlsx")
    if os.path.exists(path):
        os.remove(path)


def detect_columns(headers):
    mapping = {}
    used_indices = set()
    for field_name, aliases in COLUMN_ALIASES.items():
        for col_idx, header in enumerate(headers):
            if col_idx in used_indices:
                continue
            normalized = header.strip().lower()
            if normalized in aliases:
                mapping[field_name] = col_idx
                used_indices.add(col_idx)
                break
    return mapping


def parse_excel(file_bytes, sheet_name=None):
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = wb.sheetnames

    ws = wb[sheet_name] if sheet_name and sheet_name in sheets else wb.active
    active_sheet = ws.title

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {"error": "File is empty"}

    headers = [str(cell or "").strip() for cell in rows[0]]
    data_rows = rows[1:]

    # Filter out completely empty rows
    non_empty_rows = [r for r in data_rows if any(cell is not None and str(cell).strip() for cell in r)]

    preview = []
    for row in non_empty_rows[:5]:
        preview.append([str(cell) if cell is not None else "" for cell in row])

    auto_mapping = detect_columns(headers)

    return {
        "sheets": sheets,
        "active_sheet": active_sheet,
        "headers": headers,
        "preview_rows": preview,
        "total_rows": len(non_empty_rows),
        "auto_mapping": auto_mapping,
    }


def extract_rows_with_mapping(file_bytes, sheet_name, column_mapping):
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(all_rows) < 2:
        return []

    data_rows = all_rows[1:]
    result = []
    for raw_row in data_rows:
        # Skip completely empty rows
        if not any(cell is not None and str(cell).strip() for cell in raw_row):
            continue
        row_dict = {}
        for field_name, col_idx in column_mapping.items():
            if col_idx < len(raw_row):
                val = raw_row[col_idx]
                val = str(val).strip() if val is not None else ""
                if field_name == "ProductDescription":
                    val = val[:50]
                row_dict[field_name] = val
            else:
                row_dict[field_name] = ""
        # Store raw column values for per-store price lookups
        row_dict["_raw"] = [str(cell).strip() if cell is not None else "" for cell in raw_row]
        result.append(row_dict)
    return result


def match_categories(store_ids, rows):
    unique_subcats = set()
    for row in rows:
        sub_text = row.get("SubCateName", "").strip()
        if sub_text:
            unique_subcats.add(sub_text)

    results = {}
    for store_id in store_ids:
        all_subs = get_all_subcategories(store_id)

        sub_lookup = {}
        for sub in all_subs:
            key = sub["SubCateName"].strip().lower()
            sub_lookup[key] = {
                "SubCateID": sub["SubCateID"],
                "CateID": sub["CategoryID"],
                "SubCateName": sub["SubCateName"],
                "CategoryName": sub["CategoryName"],
            }

        store_matches = {}
        for sub_text in unique_subcats:
            match = sub_lookup.get(sub_text.lower())
            if match:
                store_matches[sub_text] = {"matched": True, **match}
            else:
                store_matches[sub_text] = {
                    "matched": False,
                    "SubCateID": None,
                    "CateID": None,
                    "SubCateName": sub_text,
                    "CategoryName": None,
                }

        results[str(store_id)] = {
            "subcategories": store_matches,
            "all_subcategories": all_subs,
        }

    return results


def _get_raw_value(row, store_mappings, sid, field):
    """Get a field value from per-store column mapping via _raw, or fall back to row dict."""
    sm = store_mappings.get(sid, {})
    col_idx = sm.get(field)
    if col_idx is not None:
        raw = row.get("_raw", [])
        return raw[col_idx].strip() if col_idx < len(raw) else ""
    return row.get(field, "").strip()


def validate_batch(rows, store_ids, category_assignments, store_mappings=None):
    if store_mappings is None:
        store_mappings = {}
    all_upcs = [r.get("ProductUPC", "").strip() for r in rows if r.get("ProductUPC", "").strip()]
    all_skus = [r.get("ProductSKU", "").strip() for r in rows if r.get("ProductSKU", "").strip()]

    existing_upcs = {}
    existing_skus = {}

    for store_id in store_ids:
        if all_upcs:
            store_existing = set()
            for chunk in _chunks(all_upcs, 100):
                placeholders = ", ".join(["%s"] * len(chunk))
                found = execute_query(
                    store_id,
                    f"SELECT ProductUPC FROM Items_tbl WHERE ProductUPC IN ({placeholders})",
                    tuple(chunk),
                )
                store_existing.update(r["ProductUPC"] for r in found)
            existing_upcs[store_id] = store_existing

        if all_skus:
            store_existing = set()
            for chunk in _chunks(all_skus, 100):
                placeholders = ", ".join(["%s"] * len(chunk))
                found = execute_query(
                    store_id,
                    f"SELECT ProductSKU FROM Items_tbl WHERE ProductSKU IN ({placeholders})",
                    tuple(chunk),
                )
                store_existing.update(r["ProductSKU"] for r in found)
            existing_skus[store_id] = store_existing

    upc_counts = {}
    for r in rows:
        upc = r.get("ProductUPC", "").strip()
        if upc:
            upc_counts[upc] = upc_counts.get(upc, 0) + 1

    store_name_cache = {}
    for sid in store_ids:
        store = get_store(sid)
        store_name_cache[sid] = store["name"] if store else f"Store {sid}"

    results = []
    for idx, row in enumerate(rows):
        errors = []
        upc = row.get("ProductUPC", "").strip()
        sku = row.get("ProductSKU", "").strip()
        desc = row.get("ProductDescription", "").strip()

        if not upc:
            errors.append({"field": "ProductUPC", "error": "UPC is required"})
        if not desc:
            errors.append({"field": "ProductDescription", "error": "Description is required"})

        if upc and len(upc) > 20:
            errors.append({"field": "ProductUPC", "error": "UPC exceeds 20 characters"})
        if sku and len(sku) > 20:
            errors.append({"field": "ProductSKU", "error": "SKU exceeds 20 characters"})

        if upc and upc_counts.get(upc, 0) > 1:
            errors.append({"field": "ProductUPC", "error": "Duplicate UPC within this import file"})

        is_duplicate = False
        if upc:
            stores_with_upc = [
                store_name_cache[sid]
                for sid in store_ids
                if upc in existing_upcs.get(sid, set())
            ]
            if stores_with_upc:
                errors.append({
                    "field": "ProductUPC",
                    "error": f"UPC already exists in: {', '.join(stores_with_upc)}",
                })
                is_duplicate = True

        if sku:
            stores_with_sku = [
                store_name_cache[sid]
                for sid in store_ids
                if sku in existing_skus.get(sid, set())
            ]
            if stores_with_sku:
                errors.append({
                    "field": "ProductSKU",
                    "error": f"SKU already exists in: {', '.join(stores_with_sku)}",
                })

        for store_id in store_ids:
            sid = str(store_id)
            sub_text = _get_raw_value(row, store_mappings, sid, "SubCateName")
            if sub_text:
                store_cats = category_assignments.get(sid, {})
                sub_match = store_cats.get(sub_text, {})
                if not sub_match.get("SubCateID"):
                    errors.append({
                        "field": "SubCateID",
                        "error": f"Subcategory '{sub_text}' not matched for {store_name_cache[store_id]}",
                    })

        if is_duplicate:
            status = "duplicate"
        elif errors:
            status = "error"
        else:
            status = "valid"

        results.append({
            "row_index": idx,
            "status": status,
            "errors": errors,
            "upc": upc,
            "sku": sku,
            "description": desc,
        })

    return results


def _progress_path(batch_id):
    return os.path.join(UPLOAD_DIR, f"import_progress_{batch_id}.json")


def _save_progress(batch_id, progress):
    path = _progress_path(batch_id)
    with _progress_lock:
        with open(path, "w") as f:
            _json.dump(progress, f)


def _load_progress(batch_id):
    path = _progress_path(batch_id)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r") as f:
            return _json.load(f)
    except Exception:
        return None


def get_import_status(batch_id):
    return _load_progress(batch_id)


def start_import(app, rows, store_ids, category_assignments, price_mode, store_mappings, skip_rows, upload_id):
    batch_id = str(uuid.uuid4())
    importable = len([i for i in range(len(rows)) if i not in set(skip_rows or [])])
    total_ops = importable * len(store_ids)
    _save_progress(batch_id, {
        "status": "running",
        "total_rows": len(rows),
        "total_stores": len(store_ids),
        "total": total_ops,
        "processed": 0,
        "succeeded": 0,
        "failed": 0,
        "skipped": len(rows) - importable,
        "results": [],
    })

    def run():
        with app.app_context():
            _execute_import(batch_id, rows, store_ids, category_assignments,
                            price_mode, store_mappings, skip_rows, upload_id)

    thread = threading.Thread(target=run, daemon=True)
    thread.start()
    return batch_id


def _execute_import(batch_id, rows, store_ids, category_assignments, price_mode, store_mappings, skip_rows, upload_id):
    progress = _load_progress(batch_id) or {}
    skip_set = set(skip_rows) if skip_rows else set()

    price_field_names = {"UnitCost", "UnitPrice", "UnitPriceA", "UnitPriceB", "UnitPriceC", "MSRPrice"}
    any_store_uses_formula = any(
        price_mode.get(str(sid), "formula") == "formula"
        for sid in store_ids
    )

    def _update_progress():
        _save_progress(batch_id, progress)

    for idx, row in enumerate(rows):
        if idx in skip_set:
            continue

        per_store_fields = {}
        for store_id in store_ids:
            sid = str(store_id)
            store_fields = {}
            sm = store_mappings.get(sid, {})
            raw = row.get("_raw", [])

            sub_text = _get_raw_value(row, store_mappings, sid, "SubCateName")
            if sub_text:
                cat_data = category_assignments.get(sid, {}).get(sub_text, {})
                if cat_data.get("SubCateID"):
                    store_fields["CateID"] = cat_data["CateID"]
                    store_fields["SubCateID"] = cat_data["SubCateID"]

            store_price_mode = price_mode.get(sid, "formula")
            if store_price_mode == "excel":
                for price_field in ("UnitCost", "UnitPrice", "UnitPriceA", "UnitPriceB", "UnitPriceC", "MSRPrice"):
                    col_idx = sm.get(price_field)
                    if col_idx is not None:
                        try:
                            val = raw[col_idx] if col_idx < len(raw) else ""
                        except IndexError:
                            val = ""
                        if val is not None and val != "":
                            try:
                                store_fields[price_field] = float(val)
                            except (ValueError, TypeError):
                                pass
                unit_price = store_fields.get("UnitPrice")
                if unit_price is not None:
                    for pf in ("UnitPriceA", "UnitPriceB", "UnitPriceC", "MSRPrice"):
                        if pf not in store_fields:
                            store_fields[pf] = unit_price

            if store_fields:
                per_store_fields[sid] = store_fields

        common_fields = {}
        skip_keys = {"SubCateName", "CateID", "SubCateID", "_raw"}
        for key, val in row.items():
            if key in skip_keys:
                continue
            if key.startswith("_"):
                continue
            # Skip price fields from row dict — they come from per-store mappings
            if key in price_field_names:
                continue
            if val is not None and val != "":
                common_fields[key] = val

        # For required-field validation to pass, add price placeholders to
        # common_fields from the first store's per-store prices
        first_sid = str(store_ids[0])
        first_store_prices = per_store_fields.get(first_sid, {})
        for pf in ("UnitCost", "UnitPrice"):
            if pf not in common_fields and pf in first_store_prices:
                common_fields[pf] = first_store_prices[pf]

        data = {
            "store_ids": store_ids,
            "common_fields": common_fields,
            "per_store_fields": per_store_fields,
            "batch_id": batch_id,
        }

        try:
            result = insert_item(data)
        except Exception as e:
            progress["failed"] += len(store_ids)
            progress["processed"] += len(store_ids)
            progress["results"].append({
                "row_index": idx,
                "upc": row.get("ProductUPC", ""),
                "description": row.get("ProductDescription", ""),
                "status": "failed",
                "errors": [{"field": "general", "error": str(e)}],
                "results": [],
            })
            _update_progress()
            continue

        store_results = result.get("results", [])
        store_succeeded = sum(1 for r in store_results if r.get("success"))
        store_failed = len(store_results) - store_succeeded
        progress["succeeded"] += store_succeeded
        progress["failed"] += store_failed
        progress["processed"] += len(store_results)

        row_status = "success" if result["success"] else "failed"
        progress["results"].append({
            "row_index": idx,
            "upc": row.get("ProductUPC", ""),
            "description": row.get("ProductDescription", ""),
            "status": row_status,
            "errors": result.get("errors", []),
            "results": store_results,
        })
        _update_progress()

    progress["status"] = "completed"
    _update_progress()
    cleanup_upload(upload_id)


def _chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]
